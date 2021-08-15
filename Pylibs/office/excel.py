# -*- coding: utf-8 -*-
import os

from openpyxl import load_workbook, Workbook
from openpyxl.utils.protection import hash_password
from openpyxl.workbook.protection import WorkbookProtection


basepath = os.getcwd()
path = os.path.join(basepath,'test.xlsx')
path1 = os.path.join(basepath,'test1.xlsx')


def test():
    wb = load_workbook(path)
    wb.security = WorkbookProtection(workbookPassword='123456', revisionsPassword = '123456', lockWindows = True, lockStructure = True, lockRevision = True)
    # wb.create_sheet("testSheet")
    wb.save(path)

def test1():
    wb = Workbook()
    wb.security = WorkbookProtection(workbookPassword='0000', revisionsPassword='0000', lockWindows=True,
                                     lockStructure=True, lockRevision=True)
    # wb.create_sheet("testSheet")
    wb.save(path1)

if __name__ == '__main__':
    test1()